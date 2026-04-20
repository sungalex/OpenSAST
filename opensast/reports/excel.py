"""openpyxl 기반 감리용 Excel 조치 현황 시트."""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from opensast.db import models
from opensast.mois.catalog import MOIS_ITEMS_BY_ID


def build_excel(scan: models.Scan, findings: list[models.Finding]) -> bytes:
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "진단요약"
    _fill_summary(summary_ws, scan, findings)

    details_ws = wb.create_sheet("상세결과")
    _fill_details(details_ws, findings)

    coverage_ws = wb.create_sheet("49개항목")
    _fill_coverage(coverage_ws, findings)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _fill_summary(ws, scan: models.Scan, findings: list[models.Finding]) -> None:
    header_font = Font(bold=True)
    ws.append(["항목", "값"])
    ws.append(["스캔 ID", scan.id])
    ws.append(["대상", scan.source_path])
    ws.append(["상태", scan.status])
    ws.append(["시작", str(scan.started_at)])
    ws.append(["종료", str(scan.finished_at)])
    ws.append(["탐지 건수", len(findings)])
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.font = header_font


def _fill_details(ws, findings: list[models.Finding]) -> None:
    headers = [
        "심각도",
        "MOIS ID",
        "분류",
        "CWE",
        "파일",
        "라인",
        "엔진",
        "룰",
        "메시지",
        "LLM 판정",
        "오탐확률",
        "조치방안",
    ]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for f in findings:
        triage_verdict = f.triage.verdict if f.triage else ""
        triage_prob = f.triage.fp_probability if f.triage else ""
        triage_fix = f.triage.recommended_fix if f.triage else ""
        ws.append(
            [
                f.severity,
                f.mois_id or "",
                f.category or "",
                ",".join(f.cwe_ids or []),
                f.file_path,
                f.start_line,
                f.engine,
                f.rule_id,
                f.message,
                triage_verdict,
                triage_prob,
                triage_fix or "",
            ]
        )


def _fill_coverage(ws, findings: list[models.Finding]) -> None:
    ws.append(["MOIS ID", "항목명", "분류", "심각도", "탐지 건수", "적합/부적합"])
    counts: dict[str, int] = {}
    for f in findings:
        if f.mois_id:
            counts[f.mois_id] = counts.get(f.mois_id, 0) + 1
    for item_id, item in MOIS_ITEMS_BY_ID.items():
        cnt = counts.get(item_id, 0)
        ws.append(
            [
                item_id,
                item.name_kr,
                item.category.value,
                item.severity.value,
                cnt,
                "부적합" if cnt > 0 else "적합",
            ]
        )
